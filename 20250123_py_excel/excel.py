import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment

def autofit_row_height(worksheet):
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value:
                # 1. Calculate the number of lines needed for the cell content
                lines = cell.value.count('\n') + 1

                # 2. Calculate the required height for the cell based on the default font size
                font_size = 11  # Change this value as needed
                cell_height = (font_size * 1.5) * lines

                # 3. Set the row height to fit the cell content
                if worksheet.row_dimensions[cell.row].height is None or worksheet.row_dimensions[cell.row].height < cell_height:
                    worksheet.row_dimensions[cell.row].height = cell_height


def autofit_column_width(worksheet):
    for column_cells in worksheet.columns:
        for cell in column_cells:
            if cell.value:
                # 4. Calculate the required width for the column based on the content length
                column_width = (len(str(cell.value)) * 1.2)

                # 5. Set the column width to fit the cell content
                if worksheet.column_dimensions[cell.column_letter].width is None or worksheet.column_dimensions[cell.column_letter].width < column_width:

                    worksheet.column_dimensions[cell.column_letter].width = column_width

def set_column_length_auto(ws):
    for col in ws.columns:
        new_column_length = max(len(str(cell.value)) for cell in col)
        new_column_letter = (get_column_letter(col[0].column))
        if new_column_length > 0:
            ws.column_dimensions[new_column_letter].width = new_column_length * 1.1

        # for cell in ws[new_column_letter]:
        #     cell.alignment = Alignment(horizontal='center')


def save_tables_to_excel(tables_data, output_filename):
    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Combined Table"

    # Define border style
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    current_row_offset = 0

    # Process each table
    for table_data in tables_data["tables"]:
        # Find the maximum row used in the current table
        max_row = max(cell_info["end_row"] for cell_info in table_data["cells"])

        # Process each cell in the table data
        for cell_info in table_data["cells"]:
            start_row = cell_info["start_row"] + current_row_offset
            end_row = cell_info["end_row"] + current_row_offset
            start_col = cell_info["start_col"]
            end_col = cell_info["end_col"]
            value = cell_info["value"]

            # Set the value for the top-left cell
            ws.cell(row=start_row, column=start_col, value=value)
            ws.cell(row=start_row, column=start_col).alignment = openpyxl.styles.Alignment(wrap_text=True)

            # Merge cells if necessary
            if start_row != end_row or start_col != end_col:
                start_cell = f"{get_column_letter(start_col)}{start_row}"
                end_cell = f"{get_column_letter(end_col)}{end_row}"
                ws.merge_cells(f"{start_cell}:{end_cell}")

            # Apply border to all cells in the range
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    ws.cell(row=row, column=col).border = thin_border

        # Add offset for the next table
        current_row_offset += max_row + 1  # Leave one blank row between tables

    # autofit_column_width(ws)
    set_column_length_auto(ws)
    autofit_row_height(ws)

    # Save the workbook to the specified output file
    wb.save(output_filename)

# Example usage
data = {
    "tables": [
        {
            "cells": [
                {
                    "start_row": 1,
                    "end_row": 1,
                    "start_col": 1,
                    "end_col": 1,
                    "value": "Name"
                },
                {
                    "start_row": 1,
                    "end_row": 1,
                    "start_col": 2,
                    "end_col": 3,
                    "value": "Age"
                },
                {
                    "start_row": 2,
                    "end_row": 2,
                    "start_col": 1,
                    "end_col": 3,
                    "value": "John"
                }
            ]
        },
        {
            "cells": [
                {
                    "start_row": 1,
                    "end_row": 1,
                    "start_col": 1,
                    "end_col": 1,
                    "value": "Name"
                },
                {
                    "start_row": 1,
                    "end_row": 1,
                    "start_col": 2,
                    "end_col": 3,
                    "value": "Age"
                },
                {
                    "start_row": 2,
                    "end_row": 3,
                    "start_col": 1,
                    "end_col": 3,
                    "value": "John\n\nDoe"
                }
            ]
        }
    ]
}

save_tables_to_excel(data, "output.xlsx")
